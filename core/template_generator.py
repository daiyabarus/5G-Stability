# ============================================================================
# core/template_generator.py
# ============================================================================
"""Generate detailed KPI reports using Excel template."""


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import date
from collections import defaultdict
from models.data_models import TowerReport



class TemplateGenerator:
    """Generate detailed reports from template.xlsx with multiple tower IDs per date."""
    
    TEMPLATE_RANGE = {
        'start_row': 1,
        'end_row': 13,
        'start_col': 1,
        'end_col': 8
    }
    
    CELL_MAPPING = {
        'tower_id': 'B1',
        'date': 'B4',
        'kpi_result_start': 9,
        'kpi_remark_start': 9
    }
    
    @staticmethod
    def format_date_sheet_name(dt: date) -> str:
        """Format date as 'Performance DD-MMM-YY' for sheet name."""
        return f"Performance {dt.strftime('%Y-%m-%d')}"
    
    @staticmethod
    def apply_specific_merges(ws, start_row: int):
        """
        Apply specific merged cells using string ranges.
        """
        merge_ranges = [
            f"B{start_row}:C{start_row}",
            f"B{start_row+1}:C{start_row+1}",
            f"B{start_row+2}:C{start_row+2}",
            f"B{start_row+3}:C{start_row+3}",
            f"A{start_row+6}:E{start_row+6}",
            f"F{start_row+6}:G{start_row+6}",
            f"H{start_row+6}:H{start_row+7}"
        ]
        
        for range_str in merge_ranges:
            try:
                ws.unmerge_cells(range_str)
            except:
                pass
            
            ws.merge_cells(range_str)
    
    @staticmethod
    def set_template_labels(ws, start_row: int):
        """Set fixed labels and merge cells for each template block."""
        
        # B2:C2 = "5G"
        b2_cell = ws.cell(row=start_row + 1, column=2)  # B2
        b2_cell.value = "5G"
        ws.merge_cells(f"B{start_row+1}:C{start_row+1}")
        
        # B3:C3 = "START-TIME"
        b3_cell = ws.cell(row=start_row + 2, column=2)  # B3
        b3_cell.value = "START-TIME"
        
        # F7:G7 = "XL Smart"
        f7_cell = ws.cell(row=start_row + 6, column=6)  # F7
        f7_cell.value = "XL Smart"
        
        # H7:H8 = "Remark"
        h7_cell = ws.cell(row=start_row + 6, column=8)  # H7
        h7_cell.value = "Remark"
        ws.merge_cells(f"H{start_row+6}:H{start_row+7}")
    
    @staticmethod
    def copy_template_block(ws, source_row: int, target_row: int):
        """
        Copy template block WITHOUT merged cells.
        """
        from copy import copy
        
        r = TemplateGenerator.TEMPLATE_RANGE
        
        for row_offset in range(r['end_row']):
            for col in range(r['start_col'], r['end_col'] + 1):
                source_cell = ws.cell(row=source_row + row_offset, column=col)
                target_cell = ws.cell(row=target_row + row_offset, column=col)
                
                if source_cell.value:
                    target_cell.value = source_cell.value
                
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
    
    @staticmethod
    def fill_tower_data(ws, start_row: int, report: TowerReport):
        """
        Fill tower-specific data into template block.
        """
        from openpyxl.styles import PatternFill, Font, Alignment
        
        RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        DARK_RED_TEXT = Font(color="FF9C0006")
        GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        DARK_GREEN_TEXT = Font(color="FF006100")
        CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
        
        tower_id_value = report.tower_id
        
        # B1:C1 - Tower ID (all towers)
        ws.cell(row=start_row, column=2).value = tower_id_value  # B1
        
        # A7:E7 - Tower ID (SAME value as B1, all towers)
        ws.cell(row=start_row + 6, column=1).value = tower_id_value  # A7
        
        # B4:C4 - Date
        ws.cell(row=start_row + 3, column=2).value = report.report_date.strftime('%d-%b-%y')  # B4
        
        # KPI data
        kpi_data = [
            (report.kpi_values.sgnb_add_success_rate, report.kpi_status.sgnb_add_success_rate),
            (report.kpi_values.sgnb_drop_rate, report.kpi_status.sgnb_drop_rate),
            (report.kpi_values.dl_user_throughput, report.kpi_status.dl_user_throughput),
            (report.kpi_values.ul_user_throughput, report.kpi_status.ul_user_throughput),
            (report.kpi_values.dl_spectrum_efficiency, report.kpi_status.dl_spectrum_efficiency)
        ]
        
        for idx, (value, status) in enumerate(kpi_data):
            row = start_row + TemplateGenerator.CELL_MAPPING['kpi_result_start'] - 1 + idx
            
            # G column - KPI values
            result_cell = ws.cell(row=row, column=7)
            result_cell.value = round(value, 2) if value is not None else "N/A"
            result_cell.alignment = CENTER_ALIGNMENT
            
            # H column - PASS/FAIL
            remark_cell = ws.cell(row=row, column=8)
            remark_cell.value = status
            remark_cell.alignment = CENTER_ALIGNMENT
            
            if status == "PASS":
                remark_cell.fill = GREEN_FILL
                remark_cell.font = DARK_GREEN_TEXT
            elif status == "FAIL":
                remark_cell.fill = RED_FILL
                remark_cell.font = DARK_RED_TEXT
    
    @staticmethod
    def group_reports_by_date(reports: List[TowerReport]) -> Dict[date, List[TowerReport]]:
        grouped = defaultdict(list)
        for report in reports:
            grouped[report.report_date].append(report)
        return dict(grouped)
    
    @staticmethod
    def create_detailed_report(reports: List[TowerReport], template_path: str, output_path: str):
        """
        Create detailed KPI report using template.xlsx.
        """
        wb = load_workbook(template_path)
        template_sheet_name = wb.sheetnames[0]
        
        date_groups = TemplateGenerator.group_reports_by_date(reports)
        
        for date_idx, (report_date, date_reports) in enumerate(sorted(date_groups.items())):
            sheet_name = TemplateGenerator.format_date_sheet_name(report_date)
            
            if date_idx == 0:
                ws = wb[template_sheet_name]
                ws.title = sheet_name
            else:
                template_ws = wb[wb.sheetnames[0]]
                ws = wb.copy_worksheet(template_ws)
                ws.title = sheet_name
            
            current_row = 1
            for tower_idx, report in enumerate(date_reports):
                if tower_idx > 0:
                    current_row = ws.max_row + 6  # GAP +6
                    TemplateGenerator.copy_template_block(ws, 1, current_row)
                
                # Apply merges
                TemplateGenerator.apply_specific_merges(ws, current_row)
                
                # Set fixed labels
                TemplateGenerator.set_template_labels(ws, current_row)
                
                # Fill tower data
                TemplateGenerator.fill_tower_data(ws, current_row, report)
                
                current_row += 13
        
        # Summary sheet (sama seperti original)
        from core.excel_generator import ExcelGenerator
        
        summary_ws = wb.create_sheet("Summary", 0)
        headers = [
            "TOWERID", "DATE", "STATUS",
            "SgNB Add Success Rate (status)", "SgNB Drop Rate (status)",
            "DL User throughput (status)", "UL User throughput (status)",
            "DL Spectrum Efficiency (status)",
            "SgNB Add Success Rate (value)", "SgNB Drop Rate (value)",
            "DL User throughput (value)", "UL User throughput (value)",
            "DL Spectrum Efficiency (value)"
        ]
        
        summary_ws.append(headers)
        
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in summary_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=14)
            cell.fill = PatternFill(fill_type="solid", fgColor="47402D")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for report in reports:
            v = report.kpi_values
            s = report.kpi_status
            
            row = [
                report.tower_id,
                report.report_date.strftime('%Y-%m-%d'),
                report.overall_status,
                s.sgnb_add_success_rate, s.sgnb_drop_rate,
                s.dl_user_throughput, s.ul_user_throughput,
                s.dl_spectrum_efficiency,
                round(v.sgnb_add_success_rate, 2) if v.sgnb_add_success_rate else "",
                round(v.sgnb_drop_rate, 2) if v.sgnb_drop_rate else "",
                round(v.dl_user_throughput, 2) if v.dl_user_throughput else "",
                round(v.ul_user_throughput, 2) if v.ul_user_throughput else "",
                round(v.dl_spectrum_efficiency, 2) if v.dl_spectrum_efficiency else ""
            ]
            summary_ws.append(row)
        
        for row_idx, report in enumerate(reports, start=2):
            status_cell = summary_ws.cell(row=row_idx, column=3)
            ExcelGenerator._format_pass_fail(status_cell, report.overall_status == "PASS")
            
            kpi_statuses = [
                (4, report.kpi_status.sgnb_add_success_rate),
                (5, report.kpi_status.sgnb_drop_rate),
                (6, report.kpi_status.dl_user_throughput),
                (7, report.kpi_status.ul_user_throughput),
                (8, report.kpi_status.dl_spectrum_efficiency)
            ]
            
            for col_idx, status in kpi_statuses:
                cell = summary_ws.cell(row=row_idx, column=col_idx)
                ExcelGenerator._format_pass_fail(cell, status == "PASS")
            
            for col_idx in range(1, 14):
                cell = summary_ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for column in summary_ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(output_path)
