# ============================================================================
# core/excel_generator.py
# ============================================================================
"""Generate Excel report with openpyxl styling."""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from typing import List
from models.data_models import TowerReport


class ExcelGenerator:
    """Create styled Excel reports with color-coded PASS/FAIL cells."""

    RED_FILL = PatternFill(
        start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"
    )
    DARK_RED_TEXT = Font(color="FF9C0006")

    GREEN_FILL = PatternFill(
        start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"
    )
    DARK_GREEN_TEXT = Font(color="FF006100")

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="47402D")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=28)
    HEADER_FONT_SMALL = Font(bold=True, color="000000", size=16)

    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _format_pass_fail(cell, is_pass: bool):
        """Format cell as PASS (green) or FAIL (red) without borders"""
        if is_pass:
            cell.fill = ExcelGenerator.GREEN_FILL
            cell.font = ExcelGenerator.DARK_GREEN_TEXT
            cell.value = "PASS"
        else:
            cell.fill = ExcelGenerator.RED_FILL
            cell.font = ExcelGenerator.DARK_RED_TEXT
            cell.value = "FAIL"

        cell.alignment = ExcelGenerator.CENTER_ALIGNMENT

    @staticmethod
    def create_report(reports: List[TowerReport], output_path: str) -> None:
        """
        Create Excel report from tower reports.

        Report Structure:
        - Sheet: "Summary"
        - Header row: bold, dark brown background, white text
        - PASS cells: green fill with dark green text
        - FAIL cells: red fill with dark red text
        - No borders on any cells
        - Auto-adjusted column widths

        Columns (new order):
        1. TOWERID
        2. DATE
        3. STATUS (overall)
        4. SgNB Add Success Rate (status)
        5. SgNB Drop Rate (status)
        6. DL User throughput (status)
        7. UL User throughput (status)
        8. DL Spectrum Efficiency (status)
        9. SgNB Add Success Rate (value)
        10. SgNB Drop Rate (value)
        11. DL User throughput (value)
        12. UL User throughput (value)
        13. DL Spectrum Efficiency (value)

        Args:
            reports: List of TowerReport objects to write
            output_path: Path to save Excel file
        """
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Define column headers in the new order
        headers = [
            "TOWERID",
            "DATE",
            "STATUS",
            "SgNB Add Success Rate (status)",
            "SgNB Drop Rate (status)",
            "DL User throughput (status)",
            "UL User throughput (status)",
            "DL Spectrum Efficiency (status)",
            "SgNB Add Success Rate (value)",
            "SgNB Drop Rate (value)",
            "DL User throughput (value)",
            "UL User throughput (value)",
            "DL Spectrum Efficiency (value)",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=14)
            cell.fill = ExcelGenerator.HEADER_FILL
            cell.alignment = ExcelGenerator.CENTER_ALIGNMENT

        for report in reports:
            v = report.kpi_values
            s = report.kpi_status

            row = [
                report.tower_id,
                report.report_date.strftime("%m/%d/%Y"),
                report.overall_status,
                s.sgnb_add_success_rate,
                s.sgnb_drop_rate,
                s.dl_user_throughput,
                s.ul_user_throughput,
                s.dl_spectrum_efficiency,
                round(v.sgnb_add_success_rate, 2) if v.sgnb_add_success_rate else "",
                round(v.sgnb_drop_rate, 2) if v.sgnb_drop_rate else "",
                round(v.dl_user_throughput, 2) if v.dl_user_throughput else "",
                round(v.ul_user_throughput, 2) if v.ul_user_throughput else "",
                round(v.dl_spectrum_efficiency, 2) if v.dl_spectrum_efficiency else "",
            ]
            ws.append(row)

        for row_idx, report in enumerate(reports, start=2):
            status_cell = ws.cell(row=row_idx, column=3)
            ExcelGenerator._format_pass_fail(
                status_cell, report.overall_status == "PASS"
            )

            kpi_statuses = [
                (4, report.kpi_status.sgnb_add_success_rate),
                (5, report.kpi_status.sgnb_drop_rate),
                (6, report.kpi_status.dl_user_throughput),
                (7, report.kpi_status.ul_user_throughput),
                (8, report.kpi_status.dl_spectrum_efficiency),
            ]

            for col_idx, status in kpi_statuses:
                cell = ws.cell(row=row_idx, column=col_idx)
                ExcelGenerator._format_pass_fail(cell, status == "PASS")

            for col_idx in range(9, 14):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = ExcelGenerator.CENTER_ALIGNMENT

            for col_idx in range(1, 3):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = ExcelGenerator.CENTER_ALIGNMENT

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(output_path)
